from django.shortcuts import render, redirect
from django.http import HttpResponse, HttpResponseRedirect
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import connection
#from braces.views import LoginRequiredMixin
from django.views.generic import CreateView, UpdateView, DetailView, DeleteView, ListView, TemplateView
from django.urls import reverse_lazy
from django.shortcuts import get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render_to_response
import request
from django.db.models import Q
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from openpyxl import Workbook
from django.contrib import messages

#librería matplotlib que permite generar gráficos
#from matplotlib import pyplot

#LIBRERIAS PARA CREACION DE DOCUMENTOS
import os
import io
from io import BytesIO
from dateutil import relativedelta as rdelta 
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, legal, A4
from reportlab.lib import colors
from reportlab.lib.units import inch, cm
from reportlab.graphics.shapes import Drawing
from reportlab.platypus import Paragraph, Table, TableStyle, Image, Spacer, SimpleDocTemplate, Frame, PageTemplate, NextPageTemplate, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY, TA_RIGHT
from reportlab.lib import colors 
from reportlab.lib.colors import white, black
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle, PropertySet
from reportlab.platypus import BaseDocTemplate, Frame, PageTemplate, Paragraph , Preformatted
from django.http import FileResponse
from django.views.generic import View
from django.core.files.storage import FileSystemStorage
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase.pdfmetrics import registerFont, registerFontFamily 
from reportlab.platypus import SimpleDocTemplate, Paragraph, TableStyle  
from reportlab.lib.styles import getSampleStyleSheet  
from reportlab.lib import colors  
from reportlab.lib.pagesizes import letter  
from reportlab.platypus import Table  

import random
from datetime import date
from datetime import datetime
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login
from .forms import ExtendedUserCreationForm, PersonaForm, AlumnoForm
from .models import Alumno, Persona
from django.contrib.auth.models import AbstractUser, User
from Curso.models import Curso, Asignatura_alumnos, Evaluacion, Archivo, Evaluacion_alumnos, Asignatura, Grupos_Alumnos, Grupos, Validacion_Coevaluacion

from Rubrica.models import Rubrica,Evaluar_Alumnos_Coevaluacion,Calificacion_Coevaluacion,Calificacion_aspecto,Aspectos_Coevaluacion,Descripcion_Aspectos_Coevaluacion
from Rubrica.forms import Evaluar_Alumnos_CoevaluacionForm
from Carrera.models import Carrera

from Curso.forms import Archivo_form

import openpyxl
from openpyxl import Workbook #para creación excel
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl import load_workbook
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.datavalidation import DataValidation
from django.forms import formset_factory, modelformset_factory, modelform_factory
# Create your views here.

def register(request):
	if request.method == 'POST':
		form = ExtendedUserCreationForm(request.POST)
		persona_form = PersonaForm(request.POST)

		if form.is_valid() and persona_form.is_valid():
			user = form.save()
			persona = persona_form.save(commit=False)
			persona.user = user
			persona.save()

			username = form.cleaned_data.get('username')
			password = form.cleaned_data.get('password1')
			user = authenticate(username=username, password=password)
			login(request,user)

			return redirect('home')

	else:
		form = ExtendedUserCreationForm()
		persona_form = PersonaForm()


	context = {'form':form, 'persona_form':persona_form}

	return render(request, 'registration/prueba.html',context)


@login_required()
def cargaManualAlumno(request,pka):

    alumno = Alumno.objects.filter(user=request.user.id)

    if request.method == "POST":
        form = AlumnoForm(request.POST)
        if form.is_valid():
            alumno = form.save(commit=False)
            alumno.user = User.objects.get(pk=request.user.id)
            alumno.save()
            return redirect('agregarAlumnoManual',pka)
    else:
            form = AlumnoForm()

    #TIPO DE USUARIO
    current_user = request.user
    usuario = Persona.objects.get(user=current_user.id)
    return render (request, 'Alumno/ingresoAlumnoManual.html',{'form':form,'alumno':alumno,'pka':pka,'usuario':usuario})


@login_required()
def editarAlumno(request,pk,ide):
    curso = Curso.objects.get(pk=ide)
    alumno = Alumno.objects.get(pk=pk)
    form = AlumnoForm(request.POST or None, instance=alumno)
    if form.is_valid():
        form.save()
        messages.success(request, 'Se ha editado la información correctamente')
        return redirect('agregarAlumnoManual',ide)

    #TIPO DE USUARIO
    current_user = request.user
    usuario = Persona.objects.get(user=current_user.id)
    return render(request, 'Alumno/editarAlumno.html', {'form': form, 'pk':pk,'ide':ide,'usuario':usuario}) 


@login_required()
def listadoAlumnosRegistrados(request,pka):
    alumnos = Alumno.objects.filter(user=request.user.id)
    alumno_en_curso = Asignatura_alumnos.objects.filter(alumno__user=request.user.id, curso=pka)

    #TIPO DE USUARIO
    current_user = request.user
    usuario = Persona.objects.get(user=current_user.id)
    return render (request, 'Alumno/listadoAlumnosRegistrados.html',{'alumnos':alumnos,'pka':pka,'alumno_en_curso':alumno_en_curso,'usuario':usuario})


@login_required()
def fichaAcademica(request,pk,pka):

    alumno = Alumno.objects.get(pk=pk)
    evaluaciones = Evaluacion.objects.filter(curso=pka)
    curso = Curso.objects.get(pk=pka)
    notas = Evaluacion_alumnos.objects.filter(curso=pka, alumno=pk)

    #TIPO DE USUARIO
    current_user = request.user
    usuario = Persona.objects.get(user=current_user.id)
    return render (request,'Alumno/fichaAcademica.html', {'notas':notas,'alumno':alumno,'evaluaciones':evaluaciones,'curso':curso,'pka':pka,'usuario':usuario})

#FUNCIÓN QUE DESCARGA LA PLANTILLA QUE SERÁ LLENADA PARA CARGAR A LOS ALUMNOS

class plantilla_alumnos_excel(TemplateView):
    def get(self,request,pk,*args,**kwargs):

    	curso = Curso.objects.get(pk=pk,profesor=request.user.id)
    	sigla_curso = curso.asignatura.sigla
    	nombre_curso = curso.asignatura.nombre
    	nombre_carrera = curso.asignatura.carrera.nombre

    	wb = Workbook()
    	ws = wb.active
    	ws['A1'] = 'Curso'
    	ws['A2'] = 'Aula Virtual'
    	ws['A3'] = 'Carrera'

    	ws.merge_cells('F5')
    	cont = 6
    	ws['A5'] = 'Rut'
    	ws['B5'] = 'Nombres'
    	ws['C5'] = 'Apellido Paterno'
    	ws['D5'] = 'Apellido Materno'
    	ws['E5'] = 'Correo PUCV'

    	ws.row_dimensions[1].height = 23

    	ws.column_dimensions['A'].width = 20
    	ws.column_dimensions['B'].width = 35
    	ws.column_dimensions['C'].width = 25
    	ws.column_dimensions['D'].width = 25
    	ws.column_dimensions['E'].width = 30


    	ws['A5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['A5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['A5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['A5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['B5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['B5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['B5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['B5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['C5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['C5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['C5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['C5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['D5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['D5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['D5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['D5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['E5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['E5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['E5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['E5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['B1'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['B1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['B1'].font = Font(name= 'Calibro', size = 8, bold=True)

    	ws['B3'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['B3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['B3'].font = Font(name= 'Calibro', size = 8, bold=True)

    	ws.cell(row = 1, column = 2).value = str(sigla_curso + ' - ' + nombre_curso)
    	ws.cell(row = 3, column = 2).value = nombre_carrera

    	i = random.randrange(9999999999999)
    	nombre_archivo = 'plantilla-subir-listado-alumnos' + str(i) + '.xlsx'
    	response = HttpResponse(content_type = "application/ms-excel")
    	content = "attachment; filename = {0}".format(nombre_archivo)
    	response['Content-Disposition'] = content
    	wb.save(response)
    	return response


@login_required()
def subir_plantilla_alumnos(request,pk):

    curso = Curso.objects.get(pk=pk)

    if request.method == 'POST':
        form = Archivo_form(request.POST, request.FILES)
        if form.is_valid():
            archivo = form.save(commit=False)
            archivo.carrera = curso.asignatura.carrera.pk
            archivo.curso = curso
            archivo.save()
            return redirect('cargar_plantilla',pk)
    else:
        form = Archivo_form()

    #TIPO DE USUARIO
    current_user = request.user
    usuario = Persona.objects.get(user=current_user.id)
    return render (request,'Alumno/subir_plantilla_alumnos.html',{'pk':pk,'form':form,'usuario':usuario})


# CARGAR DATOS PLANTILLA INGRESADA

@login_required()
def cargar_plantilla(request,pk):
    
    archivo = Archivo.objects.get(curso=pk) #obtengo el archivo subido
    asignatura = archivo.curso.asignatura.pk
    carrera = archivo.curso.asignatura.carrera.pk
    lista_alumnos = []

    doc_name = os.path.join(str(Archivo.objects.get(curso=pk))) #obtengo el valor del nombre de la ruta donde quedó guardado el doc
    doc_name_long = 'media/'+str(doc_name) #LE DOY LA RUTA DEL ARCHIVO
    doc = load_workbook(doc_name_long) #abro el archivo
    hoja = doc.active #fijo la hoja
    ultima_fila = hoja.max_row #obtengo la última fila del excel
    
    celdas_alumno = hoja[5:ultima_fila] #obtengo la matriz de la información del alumno

    for filas in celdas_alumno:    #FUNCIÓN NO REALIZADA PARA SÓLO UN ALUMNO, TOMA MÁS DE UNO
        alumno = [celda.value for celda in filas]
        lista_alumnos.append(alumno)

    i = 0
    for alumno in lista_alumnos[1:]:

        if Alumno.objects.filter(rut=alumno[0]):
            alumno_registrado = Alumno.objects.get(rut=alumno[0])
            asignatura_alumno = Asignatura_alumnos.objects.create(alumno=Alumno.objects.get(pk=alumno_registrado.pk), curso=Curso.objects.get(pk=pk))
        else:
            a = Alumno.objects.create(carrera=Carrera.objects.get(pk=carrera), rut = alumno[0], nombre= alumno[1], apellido_pat = alumno[2], apellido_mat = alumno[3], mail = alumno[4], user=User.objects.get(pk=request.user.id))
            asignatura_alumno = Asignatura_alumnos.objects.create(alumno=Alumno.objects.get(pk=a.pk), curso=Curso.objects.get(pk=pk))
    return redirect('detalleCurso', pk)


@login_required()
def misCursosAlumnos(request):
    #TIPO DE USUARIO
    current_user = request.user
    usuario = Persona.objects.get(user=current_user.id)

    nombre = current_user.username.upper()
    apellido_pat = current_user.first_name.upper()
    apellido_mat = current_user.last_name.upper()

    tipo_usuario = usuario.tipo_usuario
    alumno = Alumno.objects.get(nombre=nombre,apellido_pat=apellido_pat, apellido_mat=apellido_mat)

    for x in Alumno.objects.filter(nombre=nombre,apellido_pat=apellido_pat, apellido_mat=apellido_mat).values('carrera'):
        for a in x.values():
            carrera = a
    curso = Curso.objects.filter(anio='2021',semestre='2',asignatura__carrera=Carrera.objects.get(pk=alumno.carrera.pk))

    asignatura = Asignatura.objects.all()
    rubricas = Rubrica.objects.all()

    return render (request, 'Alumno/misCursosAlumnos.html',{'usuario':usuario, 'tipo_usuario':tipo_usuario, 'curso':curso, 'asignatura':asignatura,'usuario':usuario})


@login_required()
def detalleCursoAlumnos(request,pk):
    suma = 0

    for x in Evaluacion.objects.filter(curso=pk):
        suma = suma + x.ponderacion

    print(suma)

    rubrica = []
    curso = Curso.objects.get(pk=pk)
    evaluaciones = Evaluacion.objects.filter(curso=pk)

    for evaluacion in Evaluacion.objects.filter(curso=pk):
        rubrica = Rubrica.objects.raw('SELECT * FROM "Rubrica_rubrica" as rubrica JOIN "Curso_evaluacion" as evaluacion ON rubrica.curso_id = evaluacion.curso_id WHERE rubrica.curso_id = ' + str(pk) + ' AND evaluacion.id = ' + str(evaluacion.pk) + ';')

    #TIPO DE USUARIO
    current_user = request.user
    usuario = Persona.objects.get(user=current_user.id)

    nombre = current_user.username.upper()
    apellido_pat = current_user.first_name.upper()
    apellido_mat = current_user.last_name.upper()

    tipo_usuario = usuario.tipo_usuario

    alumno = Alumno.objects.get(nombre=nombre,apellido_pat=apellido_pat, apellido_mat=apellido_mat)
    rut = alumno.pk

    notas = Evaluacion_alumnos.objects.filter(curso=pk, alumno=Alumno.objects.get(pk=rut))

    return render(request, 'Alumno/detalleCursoAlumnos.html', {'rut':rut,'notas':notas,'usuario':usuario,'curso':curso,'pk':pk, 'pka':pk,'evaluaciones':evaluaciones,'suma':suma,'alumno':alumno,'rubrica':rubrica})

def suma_evaluacion(alumnopk):
    alumno = Alumno.objects.get(pk=alumnopk)
    cantidad = Evaluar_Alumnos_Coevaluacion.objects.filter(alumno=alumno).count()
    print(type(cantidad))

    return (cantidad)

@login_required()
def evaluarCompañeros(request,pk,pka,ide_evaluacion):
    alumno = Alumno.objects.get(pk=pk)
    evaluacion = Evaluacion.objects.get(pk=ide_evaluacion)
    curso = Curso.objects.get(pk=pka)
    grupo = Grupos_Alumnos.objects.get(alumno=pk)
    nro_grupo = grupo.grupo.pk
    integrantes = Grupos_Alumnos.objects.filter(~Q(alumno=pk), grupo=nro_grupo).order_by('alumno')
    estado_evaluacion = Validacion_Coevaluacion.objects.filter(user=User.objects.get(pk=request.user.id), evaluacion=Evaluacion.objects.get(pk=ide_evaluacion))

    for x in Grupos_Alumnos.objects.filter(grupo=nro_grupo).values('alumno').order_by('alumno'):
        for a in x.values():
            compañero = Alumno.objects.filter(pk=a)
            estado = suma_evaluacion(a)


    #TIPO DE USUARIO
    current_user = request.user
    usuario = Persona.objects.get(user=current_user.id)
    return render (request, 'Alumno/evaluarCompañeros.html', {'estado':estado,'alumnopk':alumno.pk,'ide_evaluacion':ide_evaluacion,'estado_evaluacion':estado_evaluacion,'integrantes':integrantes,'grupo':grupo,'usuario':usuario,'alumno':alumno,'evaluacion':evaluacion, 'curso':curso,'pk':pk,'pka':pka})


@login_required()
def evalua(request,alumnopk,pka,ide_evaluacion,pk):

    aspectos = Aspectos_Coevaluacion.objects.all().values().order_by('id')
    compañero = Grupos_Alumnos.objects.get(alumno=pk)
    grupo = Grupos_Alumnos.objects.get(alumno=pk)
    nro_grupo = grupo.grupo.pk
    grupo_alumno = Grupos.objects.get(pk=nro_grupo)
    print(grupo_alumno)
    evaluacion = Evaluacion.objects.get(pk=ide_evaluacion)
    curso = Curso.objects.get(pk=pka)

    descripcion = Descripcion_Aspectos_Coevaluacion.objects.all().order_by('aspecto')
    calificacion = Calificacion_aspecto.objects.all().values().order_by('id')

    #TIPO DE USUARIO
    current_user = request.user
    usuario = Persona.objects.get(user=current_user.id)

    data = {
        'form-TOTAL_FORMS': aspectos.count(),
        'form-INITIAL_FORMS': aspectos.count(),
        'form-MAX_NUM_FORMS': aspectos.count(),
    }

    puntajes_evaluados = formset_factory(Evaluar_Alumnos_CoevaluacionForm)
    formset = puntajes_evaluados(data = data, initial = list(aspectos))

    if request.method == 'POST':
        actualiza = Validacion_Coevaluacion.objects.create(alumno=Alumno.objects.get(pk=pk), user=User.objects.get(pk=current_user.id),evaluacion=Evaluacion.objects.get(pk=ide_evaluacion),flag=True)
        formset = puntajes_evaluados(request.POST, request.FILES)
        if formset.is_valid():
            i=0
            for califica in aspectos:
                creacion_puntajes = Evaluar_Alumnos_Coevaluacion.objects.create(user=User.objects.get(pk=request.user.id),alumno=Alumno.objects.get(pk=pk),grupo=grupo,evaluacion=Evaluacion.objects.get(pk=ide_evaluacion),opinion = (formset.cleaned_data[i]).get("opinion"))
                creacion_puntajes.save()
                i = i + 1
                actualiza = Grupos_Alumnos.objects.filter(alumno=pk).update(flag=True)

            return redirect('evaluarCompañeros',alumnopk, pka,ide_evaluacion)
    else:
        hola = []
        aux = {}

        for x in range(aspectos.count()):
            aux = dict(aspectos=aspectos[x], form = formset[x])
            hola.append(aux)
         
    return render(request,'Alumno/evalua.html',{'aspectos':aspectos,'descripcion':descripcion,'hola':hola,'formset':formset,'compañero':compañero,'evaluacion':evaluacion,'curso':curso,'usuario':usuario,'calificacion':calificacion,'pk':pk,'pka':pka,'ide_evaluacion':ide_evaluacion})

def suma_puntos_coevaluacion(evaluacion,alumno):

    lista_puntaje = []
    alum= Alumno.objects.get(pk=alumno)
    eva = Evaluacion.objects.get(pk=evaluacion)
    suma = 0

    for x in Aspectos_Coevaluacion.objects.all().order_by('id'):
        aspecto = x.pk
        for y in Evaluar_Alumnos_Coevaluacion.objects.filter(alumno=alum,evaluacion=eva,aspecto=Aspectos_Coevaluacion.objects.get(pk=aspecto)):
            suma = suma + int(y.opinion.nombre)
            lista_puntaje.append(int(y.opinion.nombre))

    #FALTA DIVIDIR PUNTOS POR TOTAL EN VALIDACION_COEVALUACION.objects.filter(evaluacion,alumno)
    return (lista_puntaje)


@login_required()
def verResultadosCoevaluacion(request,curso,evaluacion,alumno):
    curso = Curso.objects.get(pk=curso)

    puntaje = suma_puntos_coevaluacion(evaluacion,alumno)
    if len(puntaje) > 1:
        puntajes = suma_puntos_coevaluacion(evaluacion,alumno)
        puntaje_total = sum(puntajes)
    else:
        puntajes = [0,0,0,0,0,0,0,0,0]
        puntaje_total = sum(puntajes)

    #TIPO DE USUARIO
    current_user = request.user
    usuario = Persona.objects.get(user=current_user.id)

    return render (request,'Alumno/verResultadosCoevaluacion.html',{'puntajes':puntajes,'puntaje_total':puntaje_total,'usuario':usuario,'curso':curso})


@login_required()
def evaluaCoevaluacion(request,pk,ide_evaluacion,pka):

    alumno = Alumno.objects.get(pk=pk)
    evaluacion = Evaluacion.objects.get(pk=ide_evaluacion)
    curso = Curso.objects.get(pk=pka)
    grupo = Grupos_Alumnos.objects.get(alumno=pk)
    nro_grupo = grupo.grupo.pk
    integrantes = Grupos_Alumnos.objects.filter(~Q(alumno=pk), grupo=nro_grupo).order_by('alumno')
    estado_evaluacion = {}

    aspectos = Aspectos_Coevaluacion.objects.all().values().order_by('id')
    descripcion = Descripcion_Aspectos_Coevaluacion.objects.all().order_by('aspecto')
    calificacion = Calificacion_aspecto.objects.all().values().order_by('id')

    lista_integrantes = []

    for ab in Grupos_Alumnos.objects.filter(~Q(alumno=pk), grupo=nro_grupo).order_by('alumno'):
        lista_integrantes.append(ab.alumno.pk)

    #TIPO DE USUARIO
    current_user = request.user
    usuario = Persona.objects.get(user=current_user.id)


    data = {
        'form-TOTAL_FORMS': integrantes.count(),
        'form-INITIAL_FORMS': integrantes.count(),
        'form-MAX_NUM_FORMS': integrantes.count(),
    }

    puntajes_evaluados = formset_factory(Evaluar_Alumnos_CoevaluacionForm,extra=integrantes.count())
    formset = puntajes_evaluados(data = data, initial = list(calificacion))
    formset2 = puntajes_evaluados(data = data, initial = list(calificacion))
    formset3 = puntajes_evaluados(data = data, initial = list(calificacion))
    formset4 = puntajes_evaluados(data = data, initial = list(calificacion))
    formset5 = puntajes_evaluados(data = data, initial = list(calificacion))
    formset6 = puntajes_evaluados(data = data, initial = list(calificacion))
    formset7 = puntajes_evaluados(data = data, initial = list(calificacion))
    formset8 = puntajes_evaluados(data = data, initial = list(calificacion))
    formset9 = puntajes_evaluados(data = data, initial = list(calificacion))

    if request.method == 'POST':
        formset = puntajes_evaluados(request.POST, request.FILES, prefix="form1")
        formset2 = puntajes_evaluados(request.POST, request.FILES, prefix="form2")
        formset3 = puntajes_evaluados(request.POST, request.FILES, prefix="form3")
        formset4 = puntajes_evaluados(request.POST, request.FILES, prefix="form4")
        formset5 = puntajes_evaluados(request.POST, request.FILES, prefix="form5")
        formset6 = puntajes_evaluados(request.POST, request.FILES, prefix="form6")
        formset7 = puntajes_evaluados(request.POST, request.FILES, prefix="form7")
        formset8 = puntajes_evaluados(request.POST, request.FILES, prefix="form8")
        formset9 = puntajes_evaluados(request.POST, request.FILES, prefix="form9")

        if all([formset.is_valid() and formset2.is_valid() and formset3.is_valid()]):
            ai=0
            bi=0
            ci=0
            di=0
            ei=0
            fi=0
            gi=0
            hi=0
            ii=0
            for califica in range(len(calificacion)):
                for form in formset:
                    creacion_puntajes = Evaluar_Alumnos_Coevaluacion.objects.create(user=User.objects.get(pk=request.user.id),alumno=Alumno.objects.get(pk=lista_integrantes[bi]),grupo=grupo,evaluacion=Evaluacion.objects.get(pk=ide_evaluacion),opinion = (formset.cleaned_data[bi]).get("opinion"))
                    creacion_puntajes.save()
                    bi = bi + 1

                for form2 in formset2:
                    creacion_puntajes2 = Evaluar_Alumnos_Coevaluacion.objects.create(user=User.objects.get(pk=request.user.id),alumno=Alumno.objects.get(pk=lista_integrantes[ai]),grupo=grupo,evaluacion=Evaluacion.objects.get(pk=ide_evaluacion),opinion = (formset2.cleaned_data[ai]).get("opinion"))
                    creacion_puntajes2.save()
                    ai = ai + 1

                for form3 in formset3:
                    creacion_puntajes3 = Evaluar_Alumnos_Coevaluacion.objects.create(user=User.objects.get(pk=request.user.id),alumno=Alumno.objects.get(pk=lista_integrantes[ci]),grupo=grupo,evaluacion=Evaluacion.objects.get(pk=ide_evaluacion),opinion = (formset3.cleaned_data[ci]).get("opinion"))
                    creacion_puntajes3.save()
                    ci = ci + 1

                for form4 in formset4:
                    creacion_puntajes4 = Evaluar_Alumnos_Coevaluacion.objects.create(user=User.objects.get(pk=request.user.id),alumno=Alumno.objects.get(pk=lista_integrantes[di]),grupo=grupo,evaluacion=Evaluacion.objects.get(pk=ide_evaluacion),opinion = (formset4.cleaned_data[di]).get("opinion"))
                    creacion_puntajes4.save()
                    di = di + 1

                for form5 in formset5:
                    creacion_puntajes5 = Evaluar_Alumnos_Coevaluacion.objects.create(user=User.objects.get(pk=request.user.id),alumno=Alumno.objects.get(pk=lista_integrantes[ei]),grupo=grupo,evaluacion=Evaluacion.objects.get(pk=ide_evaluacion),opinion = (formset5.cleaned_data[ei]).get("opinion"))
                    creacion_puntajes5.save()
                    ei = ei + 1

                for form6 in formset6:
                    creacion_puntajes6 = Evaluar_Alumnos_Coevaluacion.objects.create(user=User.objects.get(pk=request.user.id),alumno=Alumno.objects.get(pk=lista_integrantes[fi]),grupo=grupo,evaluacion=Evaluacion.objects.get(pk=ide_evaluacion),opinion = (formset6.cleaned_data[fi]).get("opinion"))
                    creacion_puntajes6.save()
                    fi = fi + 1

                for form7 in formset7:
                    creacion_puntajes7 = Evaluar_Alumnos_Coevaluacion.objects.create(user=User.objects.get(pk=request.user.id),alumno=Alumno.objects.get(pk=lista_integrantes[gi]),grupo=grupo,evaluacion=Evaluacion.objects.get(pk=ide_evaluacion),opinion = (formset7.cleaned_data[gi]).get("opinion"))
                    creacion_puntajes7.save()
                    gi = gi + 1

                for form8 in formset8:
                    creacion_puntajes8 = Evaluar_Alumnos_Coevaluacion.objects.create(user=User.objects.get(pk=request.user.id),alumno=Alumno.objects.get(pk=lista_integrantes[hi]),grupo=grupo,evaluacion=Evaluacion.objects.get(pk=ide_evaluacion),opinion = (formset8.cleaned_data[hi]).get("opinion"))
                    creacion_puntajes8.save()
                    hi = hi + 1

                for form9 in formset9:
                    creacion_puntajes9 = Evaluar_Alumnos_Coevaluacion.objects.create(user=User.objects.get(pk=request.user.id),alumno=Alumno.objects.get(pk=lista_integrantes[ii]),grupo=grupo,evaluacion=Evaluacion.objects.get(pk=ide_evaluacion),opinion = (formset9.cleaned_data[ii]).get("opinion"))
                    creacion_puntajes9.save()
                    ii = ii + 1

                print('SE GUARDÓ TO')
                validacion_coevaluacion = Validacion_Coevaluacion.objects.create(flag=True,evaluacion=Evaluacion.objects.get(pk=ide_evaluacion),user=User.objects.get(pk=request.user.id))
                return redirect('evaluarCompañeros',pk, pka,ide_evaluacion)
        else:
            messages.warning(request, 'Error en el formulario ❗❗')

    else:
        formset = puntajes_evaluados(prefix="form1")
        formset2 = puntajes_evaluados(prefix="form2")
        formset3 = puntajes_evaluados(prefix="form3")
        formset4 = puntajes_evaluados(prefix="form4")
        formset5 = puntajes_evaluados(prefix="form5")
        formset6 = puntajes_evaluados(prefix="form6")
        formset7 = puntajes_evaluados(prefix="form7")
        formset8 = puntajes_evaluados(prefix="form8")
        formset9 = puntajes_evaluados(prefix="form9")

    return render (request,'Alumno/evaluaCoevaluacion.html',{'formset9':formset9,'formset8':formset8,'formset7':formset7,'formset6':formset6,'formset5':formset5,'formset4':formset4,'formset3':formset3,'formset':formset,'formset2':formset2,'usuario':usuario,'formset':formset,'aspectos':aspectos,'calificacion':calificacion,'alumno':alumno,'integrantes':integrantes,'curso':curso,'nro_grupo':nro_grupo,'evaluacion':evaluacion,'descripcion':descripcion})


@login_required()
def evaluaCoevaluacion2(request,pk,ide_evaluacion,pka):

    alumno = Alumno.objects.get(pk=pk)
    evaluacion = Evaluacion.objects.get(pk=ide_evaluacion)
    curso = Curso.objects.get(pk=pka)
    grupo = Grupos_Alumnos.objects.get(alumno=pk)
    nro_grupo = grupo.grupo.pk
    integrantes = Grupos_Alumnos.objects.filter(~Q(alumno=pk), grupo=nro_grupo).order_by('alumno')
    estado_evaluacion = {}

    aspectos = Aspectos_Coevaluacion.objects.all().values().order_by('id')
    descripcion = Descripcion_Aspectos_Coevaluacion.objects.all().order_by('aspecto')
    calificacion = Calificacion_aspecto.objects.all().values().order_by('id')
    lista_integrantes = []

    for ab in Grupos_Alumnos.objects.filter(~Q(alumno=pk), grupo=nro_grupo).order_by('alumno'):
        lista_integrantes.append(ab.alumno.pk)

    #TIPO DE USUARIO
    current_user = request.user
    usuario = Persona.objects.get(user=current_user.id)


    data = {
        'form-TOTAL_FORMS': integrantes.count(),
        'form-INITIAL_FORMS': integrantes.count(),
        'form-MAX_NUM_FORMS': integrantes.count(),
    }

    puntajes_evaluados2 = formset_factory(Evaluar_Alumnos_CoevaluacionForm)
    formset2 = puntajes_evaluados2(data = data, initial = list(calificacion))

    hola2 = []
    aux2 = {}

    for b in range(len(calificacion)):
        aux2 = dict(calificacion=calificacion[b], form = formset2[b])
        hola2.append(aux2)

    if request.method == 'POST':
        formset2 = puntajes_evaluados2(request.POST, request.FILES)

        if formset2.is_valid():
            bi=0
            for califica in range(len(calificacion)):
                creacion_puntajes2 = Evaluar_Alumnos_Coevaluacion.objects.create(user=User.objects.get(pk=request.user.id),alumno=Alumno.objects.get(pk=lista_integrantes[bi]),grupo=grupo,evaluacion=Evaluacion.objects.get(pk=ide_evaluacion),opinion = (formset2.cleaned_data[bi]).get("opinion"))
                creacion_puntajes2.save()
                bi = bi + 1
    else:

        hola2 = []
        aux2 = {}

        for b in range(len(calificacion)):
            aux2 = dict(calificacion=calificacion[b], form = formset2[b])
            hola2.append(aux2)

    return render (request,'Alumno/evaluaCoevaluacion.html',{'hola':hola,'hola2':hola2,'hola3':hola3,'hola4':hola4,'hola5':hola5,'hola6':hola6,'hola7':hola7,'hola8':hola8,'hola9':hola9,'formset':formset,'formset2':formset2,'formset3':formset3,'formset4':formset4,'formset5':formset5,'formset6':formset6,'formset7':formset7,'formset8':formset8,'formset9':formset9,'usuario':usuario,'hola':hola,'formset':formset,'aspectos':aspectos,'calificacion':calificacion,'alumno':alumno,'integrantes':integrantes,'curso':curso,'nro_grupo':nro_grupo,'evaluacion':evaluacion,'descripcion':descripcion})


