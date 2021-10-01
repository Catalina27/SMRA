from django.shortcuts import render, redirect
from django.http import HttpResponse, HttpResponseRedirect
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import connection
#from braces.views import LoginRequiredMixin
from django.views.generic import CreateView, UpdateView, DetailView, DeleteView, ListView, TemplateView
from django.urls import reverse_lazy
from django.shortcuts import get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render_to_response
import request
from django.db.models import Q
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from openpyxl import Workbook

#librería matplotlib que permite generar gráficos
#from matplotlib import pyplot

#LIBRERIAS PARA CREACION DE DOCUMENTOS
import os
import io
from io import BytesIO
from dateutil import relativedelta as rdelta 
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, legal, A4
from reportlab.lib import colors
from reportlab.lib.units import inch, cm
from reportlab.graphics.shapes import Drawing
from reportlab.platypus import Paragraph, Table, TableStyle, Image, Spacer, SimpleDocTemplate, Frame, PageTemplate, NextPageTemplate, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY, TA_RIGHT
from reportlab.lib import colors 
from reportlab.lib.colors import white, black
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle, PropertySet
from reportlab.platypus import BaseDocTemplate, Frame, PageTemplate, Paragraph , Preformatted
from django.http import FileResponse
from django.views.generic import View
from django.core.files.storage import FileSystemStorage
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase.pdfmetrics import registerFont, registerFontFamily 
from reportlab.platypus import SimpleDocTemplate, Paragraph, TableStyle  
from reportlab.lib.styles import getSampleStyleSheet  
from reportlab.lib import colors  
from reportlab.lib.pagesizes import letter  
from reportlab.platypus import Table  

import random
from datetime import date
from datetime import datetime

from .models import Curso, Asignatura, Tipo_evaluacion, Evaluacion, Evaluacion_alumnos, Asignatura_alumnos, Asignatura_alumnos, Grupos, Grupos_Alumnos, Archivo_notas
from .forms import CursoForm, AsignaturaForm, EvaluacionForm, Subir_notasform
from Usuario.models import Persona, Alumno
from django.contrib.auth.models import AbstractUser, User
from Rubrica.models import Rubrica

import openpyxl
from openpyxl import Workbook #para creación excel
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl import load_workbook
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.datavalidation import DataValidation

# Create your views here.

@login_required()
def misCursos(request):
	current_user = request.user
	usuario = Persona.objects.get(user=current_user.id)
	tipo_usuario = usuario.tipo_usuario

	curso = Curso.objects.filter(anio='2021',semestre='2')
	asignatura = Asignatura.objects.all()
	rubricas = Rubrica.objects.all()

	#TIPO DE USUARIO
	current_user = request.user
	usuario = Persona.objects.get(user=current_user.id)
	return render (request, 'Curso/misCursos.html',{'usuario':usuario, 'tipo_usuario':tipo_usuario, 'curso':curso, 'asignatura':asignatura,'usuario':usuario})


@login_required()
def agregarAsignatura(request):
	cursos = Curso.objects.filter(profesor=request.user.id)
	asignaturas = Asignatura.objects.filter(autor=request.user.id)
	cursoList = []
	user = request.user.id

	for asignaturi in Asignatura.objects.filter(autor=request.user.id):
		if Curso.objects.filter(profesor=asignaturi.autor.id).filter(asignatura=asignaturi.id):
			print('Ya está en tu listado de cursos')
		else:
			print('No está en tu listado de cursos')
			cursoList.append(asignaturi)

	if request.method == "POST":
		form = AsignaturaForm(request.POST)
		if form.is_valid():
			asignatura = form.save(commit=False)
			asignatura.autor = User.objects.get(pk=request.user.id)
			asignatura.save()
			return redirect('agregarAsignatura')
	else:
			form = AsignaturaForm()

	#TIPO DE USUARIO
	current_user = request.user
	usuario = Persona.objects.get(user=current_user.id)
	return render (request, 'Curso/agregarAsignatura.html', {'form':form, 'cursos':cursos,'cursoList':cursoList,'usuario':usuario})


@login_required()
def agregarCurso(request,pk):
	asignatura = get_object_or_404(Asignatura, pk=pk)
	print(asignatura)

	if request.method == "POST":
		form = CursoForm(request.POST or None)
		if form.is_valid():
			curso = form.save(commit=False)
			curso.asignatura = asignatura
			curso.profesor = User.objects.get(pk=request.user.id)
			curso.save()
			print('se creó')
			return redirect('misCursos')
	else:
		form = CursoForm()
		print('no se creó')

	#TIPO DE USUARIO
	current_user = request.user
	usuario = Persona.objects.get(user=current_user.id)
	return render (request, 'Curso/agregarCurso.html', {'form':form,'usuario':usuario})


@login_required()
def detalleCurso(request,pk):
    suma = 0

    for x in Evaluacion.objects.filter(curso=pk):
        suma = suma + x.ponderacion

    print(suma)

    notas = ()
    rubrica = []
    curso = Curso.objects.get(pk=pk)
    evaluaciones = Evaluacion.objects.filter(curso=pk)

    lista_notas = []
    for x in Evaluacion_alumnos.objects.filter(curso=pk).values('nota').order_by('alumno'):
        for y in x.values():
            lista_notas.append(y)

    alumno = Asignatura_alumnos.objects.filter(curso=pk)
    
    for evaluacion in Evaluacion.objects.filter(curso=pk):
        rubrica = Rubrica.objects.raw('SELECT * FROM "Rubrica_rubrica" as rubrica JOIN "Curso_evaluacion" as evaluacion ON rubrica.curso_id = evaluacion.curso_id WHERE rubrica.curso_id = ' + str(pk) + ' AND evaluacion.id = ' + str(evaluacion.pk) + ';')
        notas = Evaluacion_alumnos.objects.filter(curso=pk,evaluacion=evaluacion)   
    #TIPO DE USUARIO
    current_user = request.user
    usuario = Persona.objects.get(user=current_user.id)
    return render(request, 'Curso/detalleCurso.html', {'lista_notas':lista_notas,'notas':notas,'usuario':usuario,'curso':curso,'pk':pk, 'pka':pk,'evaluaciones':evaluaciones,'suma':suma,'alumno':alumno,'rubrica':rubrica})


@login_required()
def notas_alumno(request,pk,pka):
	
	evaluacion = Evaluacion.objects.get(pk=pk)
	curso = Curso.objects.get(pk=pka)
	notas = Evaluacion_alumnos.objects.filter(curso=pka, evaluacion=pk)
	print(pk)

	if request.method == 'POST':
		form = Subir_notasform(request.POST, request.FILES)
		if form.is_valid():
			archivo = form.save(commit=False)
			archivo.evaluacion = Evaluacion.objects.get(pk=pk)
			archivo.curso = Curso.objects.get(pk=pka)
			archivo.save()
			return redirect('cargar_plantilla_notas_alumnos',pk,pka)
	else:
		form = Subir_notasform()

	current_user = request.user
	usuario = Persona.objects.get(user=current_user.id)
	return render(request,'Curso/notas_alumno.html',{'notas':notas,'evaluacion':evaluacion,'pk':pk,'pka':pka,'form':form,'usuario':usuario})



#FUNCIÓN QUE DESCARGA LA PLANTILLA QUE SERÁ LLENADA PARA CARGAR A LOS ALUMNOS

class plantilla_notas_alumnos_excel(TemplateView):
    def get(self,request,pk,pka,*args,**kwargs):
    	current_user = request.user
    	usuario = Persona.objects.get(user=current_user.id)

    	curso = Curso.objects.get(pk=pka,profesor=current_user.id)
    	evaluacion = Evaluacion.objects.get(pk=pk)
    	alumno = Asignatura_alumnos.objects.filter(curso=curso)
    	sigla_curso = curso.asignatura.sigla
    	nombre_curso = curso.asignatura.nombre
    	nombre_carrera = curso.asignatura.carrera.nombre

    	wb = Workbook()
    	ws = wb.active
    	ws['A1'] = 'Curso'
    	ws['A2'] = 'Aula Virtual'
    	ws['A3'] = 'Carrera'

    	ws.merge_cells('F5')
    	cont = 6
    	ws['A5'] = 'Rut'
    	ws['B5'] = 'Nombres'
    	ws['C5'] = 'Apellido Paterno'
    	ws['D5'] = 'Apellido Materno'
    	ws['E5'] = 'Correo PUCV'

    	ws.row_dimensions[1].height = 23

    	ws.column_dimensions['A'].width = 20
    	ws.column_dimensions['B'].width = 35
    	ws.column_dimensions['C'].width = 25
    	ws.column_dimensions['D'].width = 25
    	ws.column_dimensions['E'].width = 30
    	ws.column_dimensions['F'].width = 40


    	ws['A5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['A5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['A5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['A5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['B5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['B5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['B5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['B5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['C5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['C5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['C5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['C5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['D5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['D5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['D5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['D5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['E5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['E5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['E5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['E5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['F5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['F5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['F5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['F5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['B1'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['B1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['B1'].font = Font(name= 'Calibro', size = 8, bold=True)

    	ws['B3'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['B3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['B3'].font = Font(name= 'Calibro', size = 8, bold=True)

    	ws.cell(row = 1, column = 2).value = str(sigla_curso + ' - ' + nombre_curso)
    	ws.cell(row = 3, column = 2).value = nombre_carrera
    	ws.cell(row = 5, column = 6).value = str(evaluacion.nombre) + str(' - %') + str(evaluacion.ponderacion)

    	for alumno in alumno:
    		ws.cell(row = cont, column = 1).value = str(alumno.alumno.rut)
    		ws.cell(row = cont, column = 2).value = alumno.alumno.nombre
    		ws.cell(row = cont, column = 3).value = alumno.alumno.apellido_pat
    		ws.cell(row = cont, column = 4).value = alumno.alumno.apellido_mat
    		ws.cell(row = cont, column = 5).value = alumno.alumno.mail
    		cont+=1

    	i = random.randrange(9999999999999)
    	nombre_archivo = 'plantilla-subir-notas-alumnos' + str(i) + '.xlsx'
    	response = HttpResponse(content_type = "application/ms-excel")
    	content = "attachment; filename = {0}".format(nombre_archivo)
    	response['Content-Disposition'] = content
    	wb.save(response)
    	return response


#FUNCIÓN QUE DESCARGA LA PLANTILLA QUE SERÁ LLENADA PARA CARGAR A LOS ALUMNOS

class plantilla_notas_grupos_excel(TemplateView):
    def get(self,request,pk,pka,*args,**kwargs):
    	current_user = request.user
    	usuario = Persona.objects.get(user=current_user.id)

    	curso = Curso.objects.get(pk=pka,profesor=current_user.id)
    	evaluacion = Evaluacion.objects.get(pk=pk)
    	alumno = Asignatura_alumnos.objects.filter(curso=curso)
    	grupos = Grupos_Alumnos.objects.filter(curso=curso).order_by('grupo')
    	sigla_curso = curso.asignatura.sigla
    	nombre_curso = curso.asignatura.nombre
    	nombre_carrera = curso.asignatura.carrera.nombre

    	wb = Workbook()
    	ws = wb.active
    	ws['A1'] = 'Curso'
    	ws['A2'] = 'Aula Virtual'
    	ws['A3'] = 'Carrera'

    	ws.merge_cells('F5')
    	cont = 6
    	ws['A5'] = 'N°'
    	ws['B5'] = 'Grupo'
    	ws['C5'] = 'Rut'
    	ws['D5'] = 'Nombres'
    	ws['E5'] = 'Apellido Paterno'
    	ws['F5'] = 'Apellido Materno'
    	ws['G5'] = 'Correo PUCV'

    	ws.row_dimensions[1].height = 23

    	ws.column_dimensions['A'].width = 5
    	ws.column_dimensions['B'].width = 10
    	ws.column_dimensions['C'].width = 25
    	ws.column_dimensions['D'].width = 35
    	ws.column_dimensions['E'].width = 20
    	ws.column_dimensions['F'].width = 20
    	ws.column_dimensions['G'].width = 40
    	ws.column_dimensions['H'].width = 40


    	ws['A5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['A5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['A5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['A5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['B5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['B5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['B5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['B5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['C5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['C5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['C5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['C5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['D5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['D5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['D5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['D5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['E5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['E5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['E5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['E5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['F5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['F5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['F5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['F5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['G5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['G5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['G5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['G5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['H5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['H5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['H5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['H5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['B1'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['B1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['B1'].font = Font(name= 'Calibro', size = 8, bold=True)

    	ws['B3'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['B3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['B3'].font = Font(name= 'Calibro', size = 8, bold=True)

    	ws.cell(row = 1, column = 2).value = str(sigla_curso + ' - ' + nombre_curso)
    	ws.cell(row = 3, column = 2).value = nombre_carrera
    	ws.cell(row = 5, column = 8).value = str(evaluacion.nombre) + str(' - %') + str(evaluacion.ponderacion)

    	for grupos in grupos:
    		ws.cell(row = cont, column = 1).value = grupos.grupo.id
    		ws.cell(row = cont, column = 2).value = grupos.grupo.nombre
    		ws.cell(row = cont, column = 3).value = str(grupos.alumno.rut)
    		ws.cell(row = cont, column = 4).value = grupos.alumno.nombre
    		ws.cell(row = cont, column = 5).value = grupos.alumno.apellido_pat
    		ws.cell(row = cont, column = 6).value = grupos.alumno.apellido_mat
    		ws.cell(row = cont, column = 7).value = grupos.alumno.mail
    		cont+=1

    	i = random.randrange(9999999999999)
    	nombre_archivo = 'plantilla-subir-notas-grupos' + str(i) + '.xlsx'
    	response = HttpResponse(content_type = "application/ms-excel")
    	content = "attachment; filename = {0}".format(nombre_archivo)
    	response['Content-Disposition'] = content
    	wb.save(response)
    	return response


# CARGAR DATOS PLANTILLA INGRESADA DE NOTAS - PARA GRUPOS E INDIVIDUALES

@login_required()
def cargar_plantilla_notas_alumnos(request,pk,pka):
    
    archivo = Archivo_notas.objects.get(curso=pka) #obtengo el archivo subido
    asignatura = archivo.curso.asignatura.pk
    carrera = archivo.curso.asignatura.carrera.pk
    evaluacion = Evaluacion.objects.get(pk=pk)
    lista_alumnos = []

    doc_name = os.path.join(str(Archivo_notas.objects.get(curso=pka))) #obtengo el valor del nombre de la ruta donde quedó guardado el doc
    doc_name_long = 'media/'+str(doc_name) #LE DOY LA RUTA DEL ARCHIVO
    doc = load_workbook(doc_name_long) #abro el archivo
    hoja = doc.active #fijo la hoja
    ultima_fila = hoja.max_row #obtengo la última fila del excel
    
    celdas_alumno = hoja[5:ultima_fila] #obtengo la matriz de la información del alumno

    for filas in celdas_alumno:    #FUNCIÓN NO REALIZADA PARA SÓLO UN ALUMNO, TOMA MÁS DE UNO
        alumno = [celda.value for celda in filas]
        lista_alumnos.append(alumno)

    cabezera = lista_alumnos[0]
    i = 0
    for alumno in lista_alumnos[1:]:
    	if evaluacion.is_grupal == False:
    		for j in range(5,len(alumno)):
    			alumno_registrado = Alumno.objects.get(rut=alumno[0])
    			evaluacion_alumno = Evaluacion_alumnos.objects.create(evaluacion = evaluacion, alumno = Alumno.objects.get(pk=alumno_registrado.pk), nota=alumno[j], curso=Curso.objects.get(pk=pka))
    	else:
    		for j in range(7,len(alumno)):
    			alumno_registrado = Alumno.objects.get(rut=alumno[2])
    			evaluacion_alumno = Evaluacion_alumnos.objects.create(evaluacion = evaluacion, alumno = Alumno.objects.get(pk=alumno_registrado.pk), nota=alumno[j], curso=Curso.objects.get(pk=pka))  
    return redirect('detalleCurso', pka)



@login_required()
def detalle_general_asignatura(request,pk):
	asignatura = Asignatura.objects.get(pk=pk)

	#TIPO DE USUARIO
	current_user = request.user
	usuario = Persona.objects.get(user=current_user.id)
	return render(request, 'Curso/detalle_general_asignatura.html', {'asignatura':asignatura,'pk':pk,'usuario':usuario})


@login_required()
def crearEvaluacion(request,pk):
	curso = Curso.objects.get(pk=pk)
	evaluaciones = Evaluacion.objects.filter(curso=pk)

	if request.method == "POST":
		form = EvaluacionForm(request.POST)
		if form.is_valid():
			evaluacion = form.save(commit=False)
			evaluacion.curso = curso
			evaluacion.save()
			return redirect('crear_evaluacion',pk)
	else:
			form = EvaluacionForm()

	#TIPO DE USUARIO
	current_user = request.user
	usuario = Persona.objects.get(user=current_user.id)
	return render (request, 'Evaluacion/crear_evaluacion.html', {'form':form, 'evaluaciones':evaluaciones,'pk':pk,'usuario':usuario})


@login_required()
def editarEvaluacion(request,pk):
	evaluacion = Evaluacion.objects.get(pk=pk)

	if request.method == 'GET':
		form = EvaluacionForm(instance=evaluacion)
	else:
		form = EvaluacionForm(request.POST, instance=evaluacion)
		if form.is_valid():
			form.save()
		return redirect('detalleCurso', evaluacion.curso.pk)

	#TIPO DE USUARIO
	current_user = request.user
	usuario = Persona.objects.get(user=current_user.id)
	return render(request, 'Evaluacion/editarEvaluacion.html', {'form':form,'evaluacion':evaluacion,'usuario':usuario})


@login_required()
def eliminarEvaluacion(request,pk):

	evaluacion = Evaluacion.objects.get(pk=pk)
	evaluacion.delete()
	
	return redirect ('detalleCurso',evaluacion.curso.pk)


####################### AGREGAR ALUMNO AL CURSO #####################

@login_required()
def agregar_alumno_a_curso(request,pk,pka):

	alumno = Alumno.objects.get(pk=pk) #obtengo el alumno
	curso = Curso.objects.get(pk=pka) #obtengo el curso

	alumno_en_curso = Asignatura_alumnos.objects.create(curso=curso, alumno=alumno)
	alumno_en_curso.save()

	return redirect('listadoAlumnosRegistrados',pka)


@login_required()
def eliminar_alumno_a_curso(request,pk,pka):

	alumno = Asignatura_alumnos.objects.get(alumno=pk) #obtengo el alumno
	alumno.delete()

	return redirect('listadoAlumnosRegistrados', pka)


#FUNCIÓN QUE DESCARGA NOTAS PARCIALES

class plantilla_descargar_notas_alumnos(TemplateView):
    def get(self,request,pk,pka,*args,**kwargs):
    	current_user = request.user
    	usuario = Persona.objects.get(user=current_user.id)

    	curso = Curso.objects.get(pk=pka,profesor=current_user.id)
    	evaluacion = Evaluacion.objects.get(pk=pk)
    	alumno = Asignatura_alumnos.objects.filter(curso=curso)
    	notas = Evaluacion_alumnos.objects.filter(curso=pka)
    	sigla_curso = curso.asignatura.sigla
    	nombre_curso = curso.asignatura.nombre
    	nombre_carrera = curso.asignatura.carrera.nombre

    	wb = Workbook()
    	ws = wb.active
    	ws['A1'] = 'Curso'
    	ws['A2'] = 'Aula Virtual'
    	ws['A3'] = 'Carrera'

    	ws.merge_cells('F5')
    	cont = 6
    	ws['A5'] = 'Rut'
    	ws['B5'] = 'Nombres'
    	ws['C5'] = 'Apellido Paterno'
    	ws['D5'] = 'Apellido Materno'
    	ws['E5'] = 'Correo PUCV'

    	ws.row_dimensions[1].height = 23

    	ws.column_dimensions['A'].width = 20
    	ws.column_dimensions['B'].width = 35
    	ws.column_dimensions['C'].width = 25
    	ws.column_dimensions['D'].width = 25
    	ws.column_dimensions['E'].width = 30
    	ws.column_dimensions['F'].width = 40


    	ws['A5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['A5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['A5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['A5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['B5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['B5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['B5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['B5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['C5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['C5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['C5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['C5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['D5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['D5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['D5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['D5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['E5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['E5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['E5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['E5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['F5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['F5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['F5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['F5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['B1'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['B1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['B1'].font = Font(name= 'Calibro', size = 8, bold=True)

    	ws['B3'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['B3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['B3'].font = Font(name= 'Calibro', size = 8, bold=True)

    	ws.cell(row = 1, column = 2).value = str(sigla_curso + ' - ' + nombre_curso)
    	ws.cell(row = 3, column = 2).value = nombre_carrera
    	ws.cell(row = 5, column = 6).value = str(evaluacion.nombre) + str(' - %') + str(evaluacion.ponderacion)

    	for notas in notas:
    		ws.cell(row = cont, column = 1).value = str(notas.alumno.rut)
    		ws.cell(row = cont, column = 2).value = notas.alumno.nombre
    		ws.cell(row = cont, column = 3).value = notas.alumno.apellido_pat
    		ws.cell(row = cont, column = 4).value = notas.alumno.apellido_mat
    		ws.cell(row = cont, column = 5).value = notas.alumno.mail
    		ws.cell(row = cont, column = 6).value = notas.nota
    		cont+=1

    	i = random.randrange(9999999999999)
    	nombre_archivo = 'notas_alumnos' + str(i) + '.xlsx'
    	response = HttpResponse(content_type = "application/ms-excel")
    	content = "attachment; filename = {0}".format(nombre_archivo)
    	response['Content-Disposition'] = content
    	wb.save(response)
    	return response


#FUNCIÓN QUE DESCARGA NOTAS FINALES

class plantilla_descargar_notas_parciales_alumnos(TemplateView):
    def get(self,request,pka,*args,**kwargs):
    	current_user = request.user
    	usuario = Persona.objects.get(user=current_user.id)

    	curso = Curso.objects.get(pk=pka,profesor=current_user.id)
    	alumno = Asignatura_alumnos.objects.filter(curso=curso)
    	notas = Evaluacion_alumnos.objects.filter(curso=pka)
    	evaluaciones = Evaluacion.objects.filter(curso=pka)
    	sigla_curso = curso.asignatura.sigla
    	nombre_curso = curso.asignatura.nombre
    	nombre_carrera = curso.asignatura.carrera.nombre

    	wb = Workbook()
    	ws = wb.active
    	ws['A1'] = 'Curso'
    	ws['A2'] = 'Aula Virtual'
    	ws['A3'] = 'Carrera'

    	ws.merge_cells('F5')
    	cont = 6
    	column_cont = 6
    	ws['A5'] = 'Rut'
    	ws['B5'] = 'Nombres'
    	ws['C5'] = 'Apellido Paterno'
    	ws['D5'] = 'Apellido Materno'
    	ws['E5'] = 'Correo PUCV'

    	ws.row_dimensions[1].height = 23

    	ws.column_dimensions['A'].width = 20
    	ws.column_dimensions['B'].width = 35
    	ws.column_dimensions['C'].width = 25
    	ws.column_dimensions['D'].width = 25
    	ws.column_dimensions['E'].width = 30
    	ws.column_dimensions['F'].width = 40


    	ws['A5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['A5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['A5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['A5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['B5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['B5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['B5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['B5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['C5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['C5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['C5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['C5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['D5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['D5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['D5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['D5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['E5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['E5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['E5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['E5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['F5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['F5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['F5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['F5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['H5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['H5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['H5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['H5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['I5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['I5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['I5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['I5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['B1'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['B1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['B1'].font = Font(name= 'Calibro', size = 8, bold=True)

    	ws['B3'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['B3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['B3'].font = Font(name= 'Calibro', size = 8, bold=True)

    	ws.cell(row = 1, column = 2).value = str(sigla_curso + ' - ' + nombre_curso)
    	ws.cell(row = 3, column = 2).value = nombre_carrera

    	for evaluaciones in evaluaciones:
    		ws.cell(row = 5, column = column_cont).value = str(evaluaciones.nombre) + str(' - %') + str(evaluaciones.ponderacion)
    		column_cont+=1 

    	for notas in notas:
    		ws.cell(row = cont, column = 1).value = str(notas.alumno.rut)
    		ws.cell(row = cont, column = 2).value = notas.alumno.nombre
    		ws.cell(row = cont, column = 3).value = notas.alumno.apellido_pat
    		ws.cell(row = cont, column = 4).value = notas.alumno.apellido_mat
    		ws.cell(row = cont, column = 5).value = notas.alumno.mail
    		ws.cell(row = cont, column = 6).value = notas.nota
    		cont+=1

    	i = random.randrange(9999999999999)
    	nombre_archivo = 'notas_parciales_alumnos' + str(i) + '.xlsx'
    	response = HttpResponse(content_type = "application/ms-excel")
    	content = "attachment; filename = {0}".format(nombre_archivo)
    	response['Content-Disposition'] = content
    	wb.save(response)
    	return response