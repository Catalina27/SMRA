from django.shortcuts import render, redirect
from django.http import HttpResponse, HttpResponseRedirect
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import connection
#from braces.views import LoginRequiredMixin
from django.views.generic import CreateView, UpdateView, DetailView, DeleteView, ListView, TemplateView
from django.urls import reverse_lazy
from django.shortcuts import get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render_to_response
import request
from django.db.models import Q
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from openpyxl import Workbook

#librería matplotlib que permite generar gráficos
#from matplotlib import pyplot

#LIBRERIAS PARA CREACION DE DOCUMENTOS
import os
import io
from io import BytesIO
from dateutil import relativedelta as rdelta 
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, legal, A4
from reportlab.lib import colors
from reportlab.lib.units import inch, cm
from reportlab.graphics.shapes import Drawing
from reportlab.platypus import Paragraph, Table, TableStyle, Image, Spacer, SimpleDocTemplate, Frame, PageTemplate, NextPageTemplate, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY, TA_RIGHT
from reportlab.lib import colors 
from reportlab.lib.colors import white, black
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle, PropertySet
from reportlab.platypus import BaseDocTemplate, Frame, PageTemplate, Paragraph , Preformatted
from django.http import FileResponse
from django.views.generic import View
from django.core.files.storage import FileSystemStorage
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase.pdfmetrics import registerFont, registerFontFamily 
from reportlab.platypus import SimpleDocTemplate, Paragraph, TableStyle  
from reportlab.lib.styles import getSampleStyleSheet  
from reportlab.lib import colors  
from reportlab.lib.pagesizes import letter  
from reportlab.platypus import Table  

import random
from datetime import date
from datetime import datetime
from datetime import date, datetime

from .models import Rubrica, Topico_rubricas, Topico, Puntaje, Calificacion_aspecto,Aplicar_RubricaGrupo,Aplicar_Rubrica,Evaluar_Alumnos_Coevaluacion
from .forms import RubricaForm, TopicoForm, PuntajeForm, Topico_rubricasForm, Calificacion_aspecto_Form,Aplicar_RubricaForm,Aplicar_RubricaGrupoForm
from django.contrib.auth.models import AbstractUser, User
from Curso.models import Evaluacion, Curso, Asignatura_alumnos, Evaluacion_alumnos, Grupos_Alumnos, Grupos, Validacion_Alumno, Validacion_Grupo
from Usuario.models import Persona, Alumno

from django.forms import formset_factory, modelformset_factory, modelform_factory
import openpyxl
from openpyxl import Workbook #para creación excel
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl import load_workbook
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.datavalidation import DataValidation


@login_required()
def misRubricas(request,pk):

	rubricas = Rubrica.objects.filter(autor=User.objects.get(pk=request.user.id))

	#TIPO DE USUARIO
	current_user = request.user
	usuario = Persona.objects.get(user=current_user.id)
	return render(request, 'Rubrica/misRubricas.html',{'pk':pk,'rubricas':rubricas,'usuario':usuario})


@login_required()
def crearRubrica(request,pk,pka):

	rubrica = Rubrica.objects.filter(evaluacion=pk)
	evaluacion = Evaluacion.objects.get(pk=pk)

	if request.method == "POST":
		form = RubricaForm(request.POST)
		if form.is_valid():
			rubrica = form.save(commit=False)
			rubrica.evaluacion = Evaluacion.objects.get(pk=pk)
			rubrica.curso = Curso.objects.get(pk=pka)
			rubrica.autor = User.objects.get(pk=request.user.id)
			rubrica.save()
			return redirect('crearRubrica',pk,pka)
	else:
			form = RubricaForm()

	#TIPO DE USUARIO
	current_user = request.user
	usuario = Persona.objects.get(user=current_user.id)
	return render (request, 'Rubrica/crear_rubrica.html', {'evaluacion':evaluacion,'form':form, 'rubrica':rubrica,'pk':pk,'pka':pka,'usuario':usuario})


@login_required()
def verRubricaEvaluacion(request,pk):

	rubrica = Rubrica.objects.get(pk=pk)

	#TIPO DE USUARIO
	current_user = request.user
	usuario = Persona.objects.get(user=current_user.id)
	return render (request, 'Rubrica/verRubricaEvaluacion.html', {'rubrica':rubrica,'pk':pk,'usuario':usuario})


@login_required()
def editarRubricaGeneral(request,pk):

	rubrica = Rubrica.objects.get(pk=pk)

	if request.method == 'GET':
		form = RubricaForm(instance=rubrica)
	else:
		form = RubricaForm(request.POST, instance=rubrica)
		if form.is_valid():
			form.save()
		return redirect('detalleCurso', rubrica.curso.pk)

	#TIPO DE USUARIO
	current_user = request.user
	usuario = Persona.objects.get(user=current_user.id)
	return render(request, 'Rubrica/editarRubricaGeneral.html', {'form':form,'rubrica':rubrica,'usuario':usuario})


@login_required()
def eliminarRubricaGeneral(request,pk):

	rubrica = Rubrica.objects.get(pk=pk)
	rubrica.delete()
	
	return redirect ('detalleCurso',rubrica.curso.pk)


@login_required()
def verRubricaGeneral(request,pk,pka):

	rubrica = Rubrica.objects.get(pk=pk)

	#TIPO DE USUARIO
	current_user = request.user
	usuario = Persona.objects.get(user=current_user.id)
	return render (request,'Rubrica/verRubricaGeneral.html',{'rubrica':rubrica,'pk':pk,'pka':pka,'usuario':usuario})


@login_required()
def llenarRubrica(request,pk,pka):

	#TIPO DE USUARIO
	current_user = request.user
	usuario = Persona.objects.get(user=current_user.id)

	curso = Curso.objects.get(pk=pka)
	rubrica = Rubrica.objects.get(pk=pk)
	aspectos = Topico_rubricas.objects.filter(topico__user=current_user.id)
	return render (request, 'Rubrica/llenarRubrica.html', {'rubrica':rubrica,'aspectos':aspectos,'pk':pk,'usuario':usuario,'pka':pka})


@login_required()
def crearAspecto(request,pk,pka):

	#TIPO DE USUARIO
	current_user = request.user
	usuario = Persona.objects.get(user=current_user.id)

	aspectos = Topico_rubricas.objects.filter(rubrica=pk)
	print(aspectos)
	rubrica = Rubrica.objects.get(pk=pk)
	pke = pk
	pka = rubrica.curso.pk

	#if Topico.objects.filter(rubrica=pk).count() > 0:
	#	rubrica = Rubrica.objects.filter(pk=pk).update(estado=3)

	if request.method == "POST":
		form = Topico_rubricasForm(request.POST)
		if form.is_valid():
			aspecto = form.save(commit=False)
			aspecto.rubrica = Rubrica.objects.get(pk=pk)
			aspecto.save()
			return redirect('crearAspecto',pk,pka)
	else:
			form = Topico_rubricasForm()

	descripcion = Puntaje.objects.filter(rubrica=pk)

	#if Topico.objects.filter(rubrica=pk,estado='Definido').count() == Topico.objects.filter(rubrica=pk).count():
	#	estado_final = True
	#else:
	#	estado_final = False

	return render (request, 'Rubrica/crearAspecto.html', {'pka':pka,'aspectos':aspectos,'descripcion':descripcion,'pke':pke,'form':form,'rubrica':rubrica,'pk':pk,'pka':pka,'usuario':usuario})

@login_required()
def eliminarAspectoRubrica(request,pk,rubrica,pka):

	aspecto = Topico_rubricas.objects.get(pk=pk)
	aspecto.delete()
	
	return redirect ('crearAspecto',rubrica,pka)



@login_required()
def modificar_puntaje_rubrica(request,pk,pka):
	rubrica = Rubrica.objects.get(pk=pk)

	calificaciones = Calificacion_aspecto.objects.all().order_by('id')

	for x in Calificacion_aspecto.objects.filter(flag=True).values('id').order_by('id'):
		for a in x.values():
			calificacion = Calificacion_aspecto.objects.get(pk=a)
			if request.method == 'GET':
				form = Calificacion_aspecto_Form(instance=calificacion)
			else:
				form = Calificacion_aspecto_Form(request.POST, instance=calificacion)
				if form.is_valid():
					form.save()
				return redirect('modificar_puntaje_rubrica', pk)

	return render (request, 'Rubrica/modificar_puntaje_rubrica.html',{'rubrica':rubrica,'calificaciones':calificaciones,'form':form,'pka':pka,'pk':pk})


@login_required()
def eliminarAspecto(request,pk):

	aspecto = Topico.objects.get(pk=pk)
	aspecto.delete()
	
	return redirect ('agregarAspectoMantenedor')


@login_required()
def puntuarAspecto(request,pke,pka,pk):

	aspecto = Topico.objects.get(pk=pk)
	rubrica = Rubrica.objects.get(pk=pke)
	calificacion = Calificacion_aspecto.objects.filter(flag=True).values()

	data = {
		'form-TOTAL_FORMS': calificacion.count(),
		'form-INITIAL_FORMS': calificacion.count(),
		'form-MAX_NUM_FORMS': calificacion.count(),
	}

	puntajes_evaluados = formset_factory(PuntajeForm)
	formset = puntajes_evaluados(data = data, initial = list(calificacion))

	if request.method == 'POST':
		formset = puntajes_evaluados(request.POST, request.FILES)
		if formset.is_valid():
			i=0
			for califica in calificacion:
				aux_2 = Calificacion_aspecto.objects.get(pk = califica["id"])
				creacion_puntajes = Puntaje.objects.create(rubrica = Rubrica.objects.get(pk=pke), descripcion = (formset.cleaned_data[i]).get("descripcion"), calificacion=aux_2,topico=Topico.objects.get(pk=pk,rubrica=pke))
				creacion_puntajes.save()
				i = i + 1
				
				if Puntaje.objects.filter(topico=pk).count() == 4:
					aspecto_update = Topico.objects.filter(pk=pk,rubrica=pke).update(estado='Definido')      
			return redirect('crearAspecto',pke)
	else:
		hola = []
		aux = {}

		for x in range(calificacion.count()):
			aux = dict(calificacion=calificacion[x], form = formset[x])
			hola.append(aux)
	#TIPO DE USUARIO
	current_user = request.user
	usuario = Persona.objects.get(user=current_user.id)


	return render (request, 'Rubrica/puntuarAspecto.html', {'hola':hola,'formset':formset,'usuario':usuario,'rubrica':rubrica,'aspecto':aspecto,'calificacion':calificacion,'pke':pke,'pka':pka,'pk':pk})


@login_required()
def rubricaFinalizada(request,pk,pka):

	aspectos = Topico_rubricas.objects.filter(rubrica=pk)
	rubrica = Rubrica.objects.get(pk=pk)

	curso = Curso.objects.get(pk=pka)
	calificacion = Calificacion_aspecto.objects.all()	
	puntajes = []

	for aspecto in Topico_rubricas.objects.filter(rubrica=pk):
		for p in Puntaje.objects.filter(topico=Topico.objects.get(pk=aspecto.topico.pk)).order_by('calificacion'):
			puntajes.append(p)

	#TIPO DE USUARIO
	current_user = request.user
	usuario = Persona.objects.get(user=current_user.id)
	return render (request, 'Rubrica/rubricaFinalizada.html',{'usuario':usuario,'rubrica':rubrica,'curso':curso,'calificacion':calificacion,'puntajes':puntajes,'aspectos':aspectos,'pk':pk,'pka':pka})


#FUNCIONES DE CREACION DE DOCUMENTOS

@login_required()
def descarga_rubrica_pdf(self,pk):

	rubrica = Rubrica.objects.get(pk=pk)
	today = date.today()
	fecha = today.strftime("%d/%m/%Y")

	response = HttpResponse(content_type='application/pdf')  
	response['content-Disposition'] = 'attachment; filename=rubrica-' + str(rubrica.evaluacion.nombre) + '.pdf'

	buffer = io.BytesIO()  
	doc = SimpleDocTemplate(buffer,  
    	pagesize=letter,  
    	rightMargin=40,  
    	leftMargin=40,  
    	topMargin=60,  
    	bottomMargin=18,  
    	)  

	sample_style_sheet = getSampleStyleSheet()

	story = []
	style = sample_style_sheet["Normal"]
	style.fontName = 'Times-Roman'
	style.fontSize = 10
	style.alignment = TA_RIGHT

    #LOGOS
	fichero_imagen = "logo_pucv.png"
	imagen_logo = Image(os.path.realpath(fichero_imagen), width=200,height=80)
	imagen_logo.align = 'LEFT'


	style.pageBreakBefore=0
	style.keepWithNext=0
	style.fontSize=11

	Aderecha = getSampleStyleSheet()
	alineadoderecha = Aderecha['Normal']
	alineadoderecha.alignment=TA_RIGHT

	parte1 = Paragraph('', style)
	parte2 = Paragraph('', alineadoderecha)


	cadena = 'FACULTAD DE INGENIERÍA'
	parrafo_1 = Paragraph(cadena, style)
	cadena_2 = 'ESCUELA DE INGENIERIA INFORMÁTICA'
	parrafo_2 = Paragraph(cadena_2, style)
	cadena_3 = '<u>' + str(rubrica.curso.asignatura.carrera.nombre) + '</u>'
	parrafo_3 = Paragraph(cadena_3, style)
	cadena_fecha = '<hr><br></br><b>Fecha: </b>'+str(fecha)
	parrafo_fecha = Paragraph(cadena_fecha,style)

	Style2 = getSampleStyleSheet()
	style_2 = Style2["Normal"]
	style_2.fontName = 'Times-Roman'
	style_2.fontSize = 15
	style_2.alignment = TA_CENTER
	cadena_4 = '<br /><b>Rúbrica <i>' + str(rubrica.evaluacion.nombre) + '</i></b> - <u>' + str(rubrica.curso.asignatura.nombre) + '</u>'
	parrafo_4 = Paragraph(cadena_4, style_2)

	cadena_salto = '<br /><br /><br /><br />'
	parrafo_salto = Paragraph(cadena_salto, style_2)

	style4 = getSampleStyleSheet()
	style_4 = style4["Normal"]
	style_4.fontName = 'Times-Roman'
	style_4.fontSize = 14
	style_4.alignment = TA_JUSTIFY
	style_4.leading = 25

    #///////////////////////////////tabla///////////////////////
	t=Table(
    	data=[
        	[[imagen_logo, parte1],[parte2]],
    	],
    	style=[
        	('VALIGN', (1, 1), (-1, -1), 'TOP'),   
    	]
    	)

	headings = ('N°','ASPECTOS','(1) INSUFICIENTE', '(2) SUFIENTE', '(3) BUENO', '(4) EXCELENTE')

	if not pk:  
		puntuacion = [(p.id, p.nombre) for p in Topico_rubricas.objects.filter(rubrica=pk) and
					  (j.descripcion) for j in Topico_rubricas.objects.filter(rubrica=pk,topico=Topico_rubricas.objects.get(pk=p))]
	else:  
		puntuacion = [(p.id, p.nombre, a.descripcion, b.descripcion,c.descripcion,d.descripcion) for p in Topico_rubricas.objects.filter(rubrica=pk)
					 for a in Puntaje.objects.filter(rubrica=pk,topico=p,calificacion=1)
					 for b in Puntaje.objects.filter(rubrica=pk,topico=p,calificacion=2)
					 for c in Puntaje.objects.filter(rubrica=pk,topico=p,calificacion=3)
					 for d in Puntaje.objects.filter(rubrica=pk,topico=p,calificacion=4)]

	ti = Table([headings] + puntuacion)  
	ti.setStyle(TableStyle(  
	[  
    	('GRID', (0, 0), (-1, -1), 1, colors.HexColor("#00030")),  
    	('LINEBELOW', (0, 0), (-1, 0), 3, colors.HexColor("#000306")),  
    	('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#29588c")),
    	('TEXTCOLOR', (0,0), (-1,-1), colors.white),
    	('TEXTCOLOR', (0,1), (15,15), colors.black)  
	]  
	)) 

	style7 = getSampleStyleSheet()
	style_7 = style4["Normal"]
	style_7.fontName = 'Times-Roman'
	style_7.fontSize = 14
	style_7.alignment = TA_CENTER
	style_7.leading = 25

	firmas = '<br /><u><i>Docente a Cargo</i></u>'
	P_firmas = Paragraph(firmas, style_7)


	styles = getSampleStyleSheet()
	styleTable = styles['BodyText']
	styleTable.alignment = TA_CENTER
	styleTable.fontSize = 10
	styleTable.fontName = 'Times-Roman'

	Presidente = Paragraph('''PAMELA HERMOSILLA MONCKTON''', styleTable)
	t2=Table(
    	data=[
        	[Presidente],

    	], colWidths=140, rowHeights=70,
    	style=[('ALIGN',(1,1),(-1,-1),'CENTER'),
    	('VALIGN',(0,0),(-1,-1),'TOP'),
        
    	] )

	story.append(t)  
	story.append(parrafo_1)
	story.append(parrafo_2)
	story.append(parrafo_3)
	story.append(parrafo_fecha)
	story.append(parrafo_4)   
	story.append(parrafo_salto)
	story.append(ti)
	story.append(parrafo_salto)
	story.append(P_firmas)    
	story.append(t2)
	doc.build(story)  
	response.write(buffer.getvalue())  
	buffer.close()  

	return response

@login_required()
def agregarAspectoMantenedor(request):

	#TIPO DE USUARIO
	current_user = request.user
	usuario = Persona.objects.get(user=current_user.id)

	if request.method == "POST":
		form = TopicoForm(request.POST)
		if form.is_valid():
			aspecto = form.save(commit=False)
			aspecto.user = User.objects.get(pk=current_user.id)
			aspecto.save()
			return redirect('agregarAspectoMantenedor')
	else:
			form = TopicoForm()

	aspectos = Topico.objects.filter(user=current_user.id)

	paginator = Paginator(aspectos,10)
	page = request.GET.get('page')
	aspectos = paginator.get_page(page)

	return render(request,'Mantenedores/agregarAspectoMantenedor.html',{'form':form,'usuario':usuario,'aspectos':aspectos})


@login_required()
def puntuarAspectoMantenedores(request,pk):

	aspecto = Topico.objects.get(pk=pk)
	calificacion = Calificacion_aspecto.objects.filter(flag=True).values()

	data = {
		'form-TOTAL_FORMS': calificacion.count(),
		'form-INITIAL_FORMS': calificacion.count(),
		'form-MAX_NUM_FORMS': calificacion.count(),
	}

	puntajes_evaluados = formset_factory(PuntajeForm)
	formset = puntajes_evaluados(data = data, initial = list(calificacion))

	if request.method == 'POST':
		formset = puntajes_evaluados(request.POST, request.FILES)
		if formset.is_valid():
			i=0
			for califica in calificacion:
				aux_2 = Calificacion_aspecto.objects.get(pk = califica["id"])
				creacion_puntajes = Puntaje.objects.create(descripcion = (formset.cleaned_data[i]).get("descripcion"), calificacion=aux_2,topico=Topico.objects.get(pk=pk))
				creacion_puntajes.save()
				i = i + 1

				if Puntaje.objects.filter(topico=pk).count() == 4:
					aspecto_update = Topico.objects.filter(pk=pk).update(estado='Definido') 

			return redirect('agregarAspectoMantenedor')
	else:
		hola = []
		aux = {}

		for x in range(calificacion.count()):
			aux = dict(calificacion=calificacion[x], form = formset[x])
			hola.append(aux)
	#TIPO DE USUARIO
	current_user = request.user
	usuario = Persona.objects.get(user=current_user.id)


	return render (request, 'Rubrica/puntuarAspecto.html', {'hola':hola,'formset':formset,'usuario':usuario,'aspecto':aspecto,'calificacion':calificacion,'pk':pk})


def calculo_puntaje(alumnopk):
	alumno = Alumno.objects.get(pk=alumnopk)
	suma = 0

	for x in Aplicar_Rubrica.objects.filter(alumno=alumno):
		suma = suma + x.calificacion.valor

	return (suma)

def calculo_puntaje_grupo(grupopk):
	grupo = Grupos.objects.get(pk=grupopk)
	suma = 0

	for x in Aplicar_RubricaGrupo.objects.filter(grupo=grupopk):
		suma = suma + x.calificacion.valor

	return (suma)


@login_required()
def aplicarRubrica(request,pk,pka):
	#TIPO DE USUARIO
	current_user = request.user
	usuario = Persona.objects.get(user=current_user.id)

	rubrica = Rubrica.objects.get(pk=pk)
	evaluacion = rubrica.evaluacion.pk
	evaluacion_1 = Evaluacion.objects.get(pk=rubrica.evaluacion.pk)
	curso = Curso.objects.get(pk=pka)
	alumnos = Asignatura_alumnos.objects.filter(curso=pka)
	grupos = Grupos.objects.all().order_by('id')
	aspectos = Topico_rubricas.objects.filter(rubrica=pk).order_by('id')
	puntaje_evaluacion = 0
	evaluado = 0
	alumno_evaluado = {}
	estado_alumno = {}
	estado_grupo = {}

	calificacion = Calificacion_aspecto.objects.all().order_by('id')	
	puntajes = []

	for aspecto in Topico_rubricas.objects.filter(rubrica=pk):
		for p in Puntaje.objects.filter(topico=Topico.objects.get(pk=aspecto.topico.pk)).order_by('calificacion'):
			puntajes.append(p)

	if evaluacion_1.is_grupal == False:
		for x in Asignatura_alumnos.objects.filter(curso=pka):
			alumnopk = x.alumno.pk
			alumno_evaluado = Aplicar_Rubrica.objects.filter(rubrica=pk,alumno=Alumno.objects.get(pk=x.alumno.pk))
			evaluado = Aplicar_Rubrica.objects.filter(rubrica=pk,alumno=Alumno.objects.get(pk=x.alumno.pk)).count()
			puntaje_evaluacion = calculo_puntaje(alumnopk)
			estado_alumno = Validacion_Alumno.objects.filter(alumno=alumnopk, evaluacion=Evaluacion.objects.get(pk=rubrica.evaluacion.pk))

	else:
		for y in Grupos.objects.all().order_by('id'):
			grupopk = y.pk
			puntaje_evaluacion = calculo_puntaje_grupo(grupopk)

	for a in Grupos.objects.all().order_by('id'):
		grupo_pk = a.pk
		for b in Validacion_Grupo.objects.filter(grupo=Grupos.objects.get(pk=grupo_pk), evaluacion=Evaluacion.objects.get(pk=rubrica.evaluacion.pk)):
			grupo_pkk = b.grupo.pk
			estado_grupo = Validacion_Grupo.objects.filter(grupo=Grupos.objects.get(pk=grupo_pk), evaluacion=Evaluacion.objects.get(pk=rubrica.evaluacion.pk))

	for c in Asignatura_alumnos.objects.filter(curso=pka):
		alumnopk = c.alumno.pk
		for d in Validacion_Alumno.objects.filter(alumno=Alumno.objects.get(pk=alumnopk), evaluacion=Evaluacion.objects.get(pk=rubrica.evaluacion.pk)):
			alumno_pk = d.alumno.pk
			estado_alumno = Validacion_Alumno.objects.filter(alumno=Alumno.objects.get(pk=alumnopk), evaluacion=Evaluacion.objects.get(pk=rubrica.evaluacion.pk))

	return render(request,'Rubrica/aplicarRubrica.html',{'usuario':usuario,'estado_alumno':estado_alumno,'estado_grupo':estado_grupo,'grupos':grupos,'evaluacion_1':evaluacion_1,'puntaje_evaluacion':puntaje_evaluacion,'evaluacion':evaluacion,'alumno_evaluado':alumno_evaluado,'evaluado':evaluado,'aspectos':aspectos,'rubrica':rubrica,'pka':pka,'pk':pk,'alumnos':alumnos,'calificacion':calificacion,'puntajes':puntajes})


@login_required()
def evaluarRubrica(request,pk,pka,alumno):

	#TIPO DE USUARIO
	current_user = request.user
	usuario = Persona.objects.get(user=current_user.id)

	rubrica = Rubrica.objects.get(pk=pk)
	evaluacion = rubrica.evaluacion.pk
	curso = Curso.objects.get(pk=pka)
	alumnos = Alumno.objects.get(pk=alumno)
	aspectos = Topico_rubricas.objects.filter(rubrica=pk).order_by('id').values()
	aspectoss = Topico_rubricas.objects.filter(rubrica=pk).order_by('id')
	lista = []

	for x in Topico_rubricas.objects.filter(rubrica=pk).order_by('id'):
		lista.append(x)

	calificacion = Calificacion_aspecto.objects.all().order_by('id')	
	puntajes = []

	for aspecto in Topico_rubricas.objects.filter(rubrica=pk):
		for p in Puntaje.objects.filter(topico=Topico.objects.get(pk=aspecto.topico.pk)).order_by('calificacion'):
			puntajes.append(p)


	data = {
		'form-TOTAL_FORMS': aspectos.count(),
		'form-INITIAL_FORMS': aspectos.count(),
		'form-MAX_NUM_FORMS': aspectos.count(),
	}

	puntajes_evaluados = formset_factory(Aplicar_RubricaForm)
	formset = puntajes_evaluados(data = data, initial = list(aspectos))

	if request.method == 'POST':
		estado_final = Validacion_Alumno.objects.create(alumno=Alumno.objects.get(pk=alumno),evaluacion=Evaluacion.objects.get(pk=evaluacion),flag=True)
		formset = puntajes_evaluados(request.POST, request.FILES)
		if formset.is_valid():
			i=0
			for califica in aspectos:
				aux_2 = Topico_rubricas.objects.get(pk = califica["id"])
				creacion_puntajes = Aplicar_Rubrica.objects.create(aspecto=Topico.objects.get(pk=aux_2.topico.pk),evaluacion=Evaluacion.objects.get(pk=evaluacion),calificacion = (formset.cleaned_data[i]).get("calificacion"), rubrica=Rubrica.objects.get(pk=pk),alumno=Alumno.objects.get(pk=alumno))
				creacion_puntajes.save()
				i = i + 1

			return redirect('aplicarRubrica',pk,pka)
	else:
		hola = []
		aux = {}

		for x in range(aspectos.count()):
			aux = dict(aspectos=aspectos[x], form = formset[x])
			hola.append(aux)
	return render (request,'Rubrica/evaluarRubrica.html',{'usuario':usuario,'lista':lista,'formset':formset,'hola':hola,'aspectoss':aspectoss,'aspectos':aspectos,'rubrica':rubrica,'pka':pka,'pk':pk,'alumnos':alumnos,'calificacion':calificacion,'puntajes':puntajes})	


@login_required()
def evaluarRubricaGrupo(request,pk,pka,grupo):

	#TIPO DE USUARIO
	current_user = request.user
	usuario = Persona.objects.get(user=current_user.id)

	rubrica = Rubrica.objects.get(pk=pk)
	evaluacion = rubrica.evaluacion.pk
	curso = Curso.objects.get(pk=pka)
	grupos = Grupos.objects.get(pk=grupo)
	alumnos = Grupos_Alumnos.objects.all().order_by('alumno')
	aspectos = Topico_rubricas.objects.filter(rubrica=pk).order_by('id').values()
	aspectoss = Topico_rubricas.objects.filter(rubrica=pk).order_by('id')
	lista = []

	for x in Topico_rubricas.objects.filter(rubrica=pk).order_by('id'):
		lista.append(x)

	calificacion = Calificacion_aspecto.objects.all().order_by('id')	
	puntajes = []

	for aspecto in Topico_rubricas.objects.filter(rubrica=pk):
		for p in Puntaje.objects.filter(topico=Topico.objects.get(pk=aspecto.topico.pk)).order_by('calificacion'):
			puntajes.append(p)


	data = {
		'form-TOTAL_FORMS': aspectos.count(),
		'form-INITIAL_FORMS': aspectos.count(),
		'form-MAX_NUM_FORMS': aspectos.count(),
	}

	puntajes_evaluados = formset_factory(Aplicar_RubricaGrupoForm)
	formset = puntajes_evaluados(data = data, initial = list(aspectos))

	if request.method == 'POST':
		grupo_update = Validacion_Grupo.objects.create(grupo=Grupos.objects.get(pk=grupo),flag=True,evaluacion=Evaluacion.objects.get(pk=evaluacion))
		formset = puntajes_evaluados(request.POST, request.FILES)
		if formset.is_valid():
			i=0
			for califica in aspectos:
				aux_2 = Topico_rubricas.objects.get(pk = califica["id"])
				creacion_puntajes = Aplicar_RubricaGrupo.objects.create(aspecto=Topico.objects.get(pk=aux_2.topico.pk),evaluacion=Evaluacion.objects.get(pk=evaluacion),calificacion = (formset.cleaned_data[i]).get("calificacion"), rubrica=Rubrica.objects.get(pk=pk),grupo=grupos)
				creacion_puntajes.save()
				i = i + 1
			return redirect('aplicarRubrica',pk,pka)
	else:
		hola = []
		aux = {}

		for x in range(aspectos.count()):
			aux = dict(aspectos=aspectos[x], form = formset[x])
			hola.append(aux)
	return render (request,'Rubrica/evaluarRubrica.html',{'usuario':usuario,'lista':lista,'formset':formset,'hola':hola,'aspectoss':aspectoss,'aspectos':aspectos,'rubrica':rubrica,'pka':pka,'pk':pk,'alumnos':alumnos,'calificacion':calificacion,'puntajes':puntajes})	


#FUNCIÓN QUE DESCARGA LA PLANTILLA QUE SERÁ LLENADA PARA CARGAR A LOS ALUMNOS

class plantilla_evaluacion_rubrica_alumnos_excel(TemplateView):
    def get(self,request,pk,pka,rubric,puntaje_evaluacion,*args,**kwargs):
    	current_user = request.user
    	usuario = Persona.objects.get(user=current_user.id)

    	curso = Curso.objects.get(pk=pka)
    	evaluacion = Evaluacion.objects.get(pk=pk)
    	alumno = Asignatura_alumnos.objects.filter(curso=curso)
    	rubrica = Rubrica.objects.get(pk=rubric)
    	notas = Evaluacion_alumnos.objects.filter(curso=pka)
    	sigla_curso = curso.asignatura.sigla
    	nombre_curso = curso.asignatura.nombre
    	nombre_carrera = curso.asignatura.carrera.nombre

    	wb = Workbook()
    	ws = wb.active
    	ws['A1'] = 'Curso'
    	ws['A2'] = 'Aula Virtual'
    	ws['A3'] = 'Carrera'

    	ws.merge_cells('F5')
    	cont = 6
    	ws['A5'] = 'Rut'
    	ws['B5'] = 'Nombres'
    	ws['C5'] = 'Apellido Paterno'
    	ws['D5'] = 'Apellido Materno'
    	ws['E5'] = 'Correo PUCV'

    	ws.row_dimensions[1].height = 23

    	ws.column_dimensions['A'].width = 20
    	ws.column_dimensions['B'].width = 35
    	ws.column_dimensions['C'].width = 25
    	ws.column_dimensions['D'].width = 25
    	ws.column_dimensions['E'].width = 30
    	ws.column_dimensions['F'].width = 40
    	ws.column_dimensions['G'].width = 40


    	ws['A5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['A5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['A5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['A5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['B5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['B5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['B5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['B5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['C5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['C5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['C5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['C5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['D5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['D5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['D5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['D5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['E5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['E5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['E5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['E5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['F5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['F5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['F5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['F5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['B1'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['B1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['B1'].font = Font(name= 'Calibro', size = 8, bold=True)

    	ws['B3'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['B3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['B3'].font = Font(name= 'Calibro', size = 8, bold=True)

    	ws.cell(row = 1, column = 2).value = str(sigla_curso + ' - ' + nombre_curso)
    	ws.cell(row = 3, column = 2).value = nombre_carrera
    	ws.cell(row = 2, column = 2).value = evaluacion.nombre
    	ws.cell(row = 5, column = 6).value = 'Puntaje Obtenido'

    	for alumno in alumno:
    		puntaje = calculo_puntaje(alumno.alumno.pk)
    		ws.cell(row = cont, column = 1).value = str(alumno.alumno.rut)
    		ws.cell(row = cont, column = 2).value = alumno.alumno.nombre
    		ws.cell(row = cont, column = 3).value = alumno.alumno.apellido_pat
    		ws.cell(row = cont, column = 4).value = alumno.alumno.apellido_mat
    		ws.cell(row = cont, column = 5).value = alumno.alumno.mail
    		ws.cell(row = cont, column = 6).value = puntaje
    		cont+=1

    	i = random.randrange(9999999999999)
    	nombre_archivo = 'puntuacion-alumnos-rubrica' + str(i) + '.xlsx'
    	response = HttpResponse(content_type = "application/ms-excel")
    	content = "attachment; filename = {0}".format(nombre_archivo)
    	response['Content-Disposition'] = content
    	wb.save(response)
    	return response



#FUNCIÓN QUE DESCARGA LA PLANTILLA QUE SERÁ LLENADA PARA CARGAR A LOS ALUMNOS

class plantilla_evaluacion_rubrica_grupos_excel(TemplateView):
    def get(self,request,pk,pka,rubric,puntaje_evaluacion,*args,**kwargs):
    	current_user = request.user
    	usuario = Persona.objects.get(user=current_user.id)

    	curso = Curso.objects.get(pk=pka)
    	grupos = Grupos_Alumnos.objects.all().order_by('grupo')
    	evaluacion = Evaluacion.objects.get(pk=pk)
    	alumno = Asignatura_alumnos.objects.filter(curso=curso)
    	rubrica = Rubrica.objects.get(pk=rubric)
    	notas = Evaluacion_alumnos.objects.filter(curso=pka)
    	sigla_curso = curso.asignatura.sigla
    	nombre_curso = curso.asignatura.nombre
    	nombre_carrera = curso.asignatura.carrera.nombre

    	wb = Workbook()
    	ws = wb.active
    	ws['A1'] = 'Curso'
    	ws['A2'] = 'Aula Virtual'
    	ws['A3'] = 'Carrera'

    	ws.merge_cells('F5')
    	cont = 6
    	ws['A5'] = 'N°'
    	ws['B5'] = 'Grupo'
    	ws['C5'] = 'Rut'
    	ws['D5'] = 'Nombres'
    	ws['E5'] = 'Apellido Paterno'
    	ws['F5'] = 'Apellido Materno'
    	ws['G5'] = 'Correo PUCV'

    	ws.row_dimensions[1].height = 23

    	ws.column_dimensions['A'].width = 5
    	ws.column_dimensions['B'].width = 10
    	ws.column_dimensions['C'].width = 25
    	ws.column_dimensions['D'].width = 35
    	ws.column_dimensions['E'].width = 20
    	ws.column_dimensions['F'].width = 20
    	ws.column_dimensions['G'].width = 40
    	ws.column_dimensions['H'].width = 40


    	ws['A5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['A5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['A5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['A5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['B5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['B5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['B5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['B5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['C5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['C5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['C5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['C5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['D5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['D5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['D5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['D5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['E5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['E5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['E5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['E5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['F5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['F5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['F5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['F5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['G5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['G5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['G5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['G5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['H5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['H5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['H5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['H5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['B1'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['B1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['B1'].font = Font(name= 'Calibro', size = 8, bold=True)

    	ws['B3'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['B3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['B3'].font = Font(name= 'Calibro', size = 8, bold=True)

    	ws.cell(row = 1, column = 2).value = str(sigla_curso + ' - ' + nombre_curso)
    	ws.cell(row = 3, column = 2).value = nombre_carrera
    	ws.cell(row = 5, column = 8).value = 'Puntaje Obtenido'

    	for grupos in grupos:
    		puntaje = calculo_puntaje_grupo(grupos.grupo.pk)    		
    		ws.cell(row = cont, column = 1).value = grupos.grupo.id
    		ws.cell(row = cont, column = 2).value = grupos.grupo.nombre
    		ws.cell(row = cont, column = 3).value = str(grupos.alumno.rut)
    		ws.cell(row = cont, column = 4).value = grupos.alumno.nombre
    		ws.cell(row = cont, column = 5).value = grupos.alumno.apellido_pat
    		ws.cell(row = cont, column = 6).value = grupos.alumno.apellido_mat
    		ws.cell(row = cont, column = 7).value = grupos.alumno.mail
    		ws.cell(row = cont, column = 8).value = puntaje
    		cont+=1

    	i = random.randrange(9999999999999)
    	nombre_archivo = 'plantilla-subir-notas-grupos' + str(i) + '.xlsx'
    	response = HttpResponse(content_type = "application/ms-excel")
    	content = "attachment; filename = {0}".format(nombre_archivo)
    	response['Content-Disposition'] = content
    	wb.save(response)
    	return response


@login_required()
def coevaluaciones(request):

	#TIPO DE USUARIO
	current_user = request.user
	usuario = Persona.objects.get(user=current_user.id)

	evaluaciones = Evaluacion.objects.all()
	curso = Curso.objects.filter(anio='2021',semestre='2')

	return render (request, 'Coevaluaciones/coevaluaciones.html',{'usuario':usuario,'evaluaciones':evaluaciones,'curso':curso})


@login_required()
def coevaluaciones_evaluaciones(request,pk):

	#TIPO DE USUARIO
	current_user = request.user
	usuario = Persona.objects.get(user=current_user.id)

	evaluaciones = Evaluacion.objects.filter(curso=pk)
	curso = Curso.objects.get(pk=pk)

	return render (request, 'Coevaluaciones/coevaluaciones_evaluaciones.html',{'pk':pk,'usuario':usuario,'evaluaciones':evaluaciones,'curso':curso})


def calculo_puntaje_promedio_coevaluacion(alumno,pka):
	alumnos = Alumno.objects.get(pk=alumno)
	evaluacion = Evaluacion.objects.get(pk=pka)
	suma = 0

	for a in Evaluar_Alumnos_Coevaluacion.objects.filter(alumno=alumno,evaluacion=pka):
		suma = suma + a.opinion.nombre

	return (suma)


#FUNCIÓN QUE DESCARGA LA PLANTILLA QUE SERÁ LLENADA PARA CARGAR A LOS ALUMNOS

class plantilla_descarga_coevaluacion_alumnos(TemplateView):
    def get(self,request,pk,pka,*args,**kwargs):
    	current_user = request.user
    	usuario = Persona.objects.get(user=current_user.id)

    	curso = Curso.objects.get(pk=pk)
    	grupos = Grupos_Alumnos.objects.all().order_by('grupo')
    	evaluacion = Evaluacion.objects.get(pk=pka)
    	alumno = Asignatura_alumnos.objects.filter(curso=curso)
    	notas = Evaluacion_alumnos.objects.filter(curso=pk)

    	sigla_curso = curso.asignatura.sigla
    	nombre_curso = curso.asignatura.nombre
    	nombre_carrera = curso.asignatura.carrera.nombre

    	wb = Workbook()
    	ws = wb.active
    	ws['A1'] = 'Curso'
    	ws['A2'] = 'Aula Virtual'
    	ws['A3'] = 'Carrera'

    	ws.merge_cells('F5')
    	cont = 6
    	ws['A5'] = 'N°'
    	ws['B5'] = 'Grupo'
    	ws['A5'] = 'Rut'
    	ws['B5'] = 'Nombres'
    	ws['C5'] = 'Apellido Paterno'
    	ws['D5'] = 'Apellido Materno'
    	ws['E5'] = 'Correo PUCV'

    	ws.row_dimensions[1].height = 23

    	ws.column_dimensions['A'].width = 10
    	ws.column_dimensions['B'].width = 10
    	ws.column_dimensions['C'].width = 25
    	ws.column_dimensions['D'].width = 35
    	ws.column_dimensions['E'].width = 20
    	ws.column_dimensions['F'].width = 20
    	ws.column_dimensions['G'].width = 40
    	ws.column_dimensions['H'].width = 45


    	ws['A5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['A5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['A5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['A5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['B5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['B5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['B5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['B5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['C5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['C5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['C5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['C5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['D5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['D5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['D5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['D5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['E5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['E5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['E5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['E5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['F5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['F5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['F5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['F5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['G5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['G5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['G5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['G5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['H5'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['H5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['H5'].fill = PatternFill(start_color = 'FCBC32', end_color = 'FCBC32', fill_type = "solid")
    	ws['H5'].font = Font(name= 'Calibro', size = 10, bold=True)

    	ws['B1'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['B1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['B1'].font = Font(name= 'Calibro', size = 8, bold=True)

    	ws['B3'].alignment = Alignment(horizontal = "center", vertical = "center")
    	ws['B3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    	ws['B3'].font = Font(name= 'Calibro', size = 8, bold=True)

    	ws.cell(row = 1, column = 2).value = str(sigla_curso + ' - ' + nombre_curso)
    	ws.cell(row = 3, column = 2).value = nombre_carrera
    	ws.cell(row = 5, column = 8).value = 'Puntaje Obtenido Coevaluacion'

    	for grupos in grupos:
    		puntaje = calculo_puntaje_promedio_coevaluacion(grupos.alumno.pk,pka)    		
    		ws.cell(row = cont, column = 1).value = grupos.grupo.id
    		ws.cell(row = cont, column = 2).value = grupos.grupo.nombre
    		ws.cell(row = cont, column = 3).value = str(grupos.alumno.rut)
    		ws.cell(row = cont, column = 4).value = grupos.alumno.nombre
    		ws.cell(row = cont, column = 5).value = grupos.alumno.apellido_pat
    		ws.cell(row = cont, column = 6).value = grupos.alumno.apellido_mat
    		ws.cell(row = cont, column = 7).value = grupos.alumno.mail
    		ws.cell(row = cont, column = 8).value = puntaje
    		cont+=1

    	i = random.randrange(9999999999999)
    	nombre_archivo = 'coevaluacion_alumnos_' + str(evaluacion.nombre) + str(i) + '.xlsx'
    	response = HttpResponse(content_type = "application/ms-excel")
    	content = "attachment; filename = {0}".format(nombre_archivo)
    	response['Content-Disposition'] = content
    	wb.save(response)
    	return response